"""
Habesha Build Hub — ADMIN BOT v3
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Key changes:
  • Full price management: add / edit / delete / bulk upload CSV+Excel
  • Category management: add / archive / reorder — all from bot, no code edits
  • Auto-route POs to matching suppliers (no manual step needed)
  • PO terminology throughout
  • Broadcast with target selection
  • Supplier verification flow
  • File upload parsing (CSV + Excel via openpyxl + csv module)
  • Lightweight audit via platform_events table
"""

import logging, csv, io, json
from telegram import Update, Bot
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)
from config import (
    ADMIN_BOT_TOKEN, BUYER_BOT_TOKEN, SUPPLIER_BOT_TOKEN, ADMIN_IDS,
    ADM_PRICE_EDIT, ADM_PRICE_ADD, ADM_CAT_ADD, ADM_BROADCAST_TEXT,
    ADM_BOQ_DELIVER_FILE, ADM_BOQ_DELIVER_NOTE,
)
from models.database import (
    get_connection, get_platform_stats, upsert_supplier,
    get_supplier, get_supplier_by_id, get_suppliers_matching_categories,
    get_buyer_by_id, get_po, get_open_pos, get_po_quotes,
    get_latest_prices, get_all_active_prices,
    add_price, edit_price, delete_price, rollback_batch, bulk_import_prices,
    get_active_categories, add_category, edit_category, archive_category,
    log_event
)
from keyboards.builders import (
    admin_menu_kb, verify_action_kb, price_mgmt_kb, cat_mgmt_kb,
    back_admin, kb, lead_kb, quotes_kb
)
from utils.messages import (
    fmt_platform_stats, fmt_po_lead, fmt_price_report, fmt_quotes
)
from utils.files import send_file_preview

logging.basicConfig(format="%(asctime)s [ADMIN] %(levelname)s %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)


# ── GUARD ──────────────────────────────────────────────────────────────────────

def admin_only(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user.id not in ADMIN_IDS:
            await update.effective_message.reply_text("⛔ Unauthorized.")
            return
        return await func(update, context)
    wrapper.__name__ = func.__name__
    return wrapper


# ── /start ─────────────────────────────────────────────────────────────────────

@admin_only
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    stats = get_platform_stats()
    await update.message.reply_text(
        f"🛠️ *HBH Admin Dashboard*\n\n{fmt_platform_stats(stats)}",
        parse_mode="Markdown", reply_markup=admin_menu_kb()
    )


# ── MAIN MENU ──────────────────────────────────────────────────────────────────

@admin_only
async def adm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    action = q.data.split(":")[1]

    # ── Stats ──────────────────────────────────────────────────────────────────
    if action in ("home", "stats"):
        stats = get_platform_stats()
        await q.edit_message_text(
            f"🛠️ *HBH Admin*\n\n{fmt_platform_stats(stats)}",
            parse_mode="Markdown", reply_markup=admin_menu_kb()
        )

    # ── Buyers ────────────────────────────────────────────────────────────────
    elif action == "buyers":
        conn = get_connection()
        rows = conn.execute(
            "SELECT * FROM buyers ORDER BY created_at DESC LIMIT 15"
        ).fetchall()
        conn.close()
        text = "👥 *Recent Buyers*\n\n"
        for b in rows:
            b = dict(b)
            text += (f"• `HBH-B-{b['buyer_id']:04d}` {b.get('name','—')} "
                     f"| {b.get('buyer_type','—')} | {b.get('city','—')}\n")
        await q.edit_message_text(text or "No buyers yet.",
                                   parse_mode="Markdown", reply_markup=admin_menu_kb())

    # ── Suppliers ─────────────────────────────────────────────────────────────
    elif action == "suppliers":
        conn = get_connection()
        rows = conn.execute(
            "SELECT * FROM suppliers ORDER BY created_at DESC LIMIT 15"
        ).fetchall()
        conn.close()
        text = "🏭 *Recent Suppliers*\n\n"
        for s in rows:
            s    = dict(s)
            v    = "✅" if s.get('verified') else "❌"
            cats = s.get('categories', '[]')
            text += (f"• `HBH-S-{s['supplier_id']:04d}` {s.get('business_name','—')} "
                     f"{v} | score {s.get('score',0):.1f}\n")
        await q.edit_message_text(text or "No suppliers yet.",
                                   parse_mode="Markdown", reply_markup=admin_menu_kb())

    # ── Open POs ──────────────────────────────────────────────────────────────
    elif action == "pos":
        pos = get_open_pos()
        if not pos:
            await q.edit_message_text("No open purchase orders.",
                                       reply_markup=admin_menu_kb()); return
        buttons = []
        text    = "📋 *Open Purchase Orders*\n\n"
        for po in pos:
            text += (f"• `{po['po_code']}` — {po.get('buyer_name','?')} "
                     f"| quotes: {po.get('quotes_count',0)}\n")
            buttons.append([(f"🚀 Route {po['po_code']}", f"adm_route:{po['po_id']}")])
            if po.get('quotes_count', 0) > 0:
                buttons.append([(f"📊 Push quotes → buyer", f"adm_pushquotes:{po['po_id']}")])
        buttons.append([("🔙 Back", "adm:home")])
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb(buttons))

    # ── BOQ Queue ─────────────────────────────────────────────────────────────
    elif action == "boqs":
        conn = get_connection()
        rows = conn.execute(
            "SELECT j.*, b.name as buyer_name FROM boq_jobs j "
            "JOIN buyers b ON j.buyer_id=b.buyer_id "
            "WHERE j.status IN ('received','reviewing') "
            "ORDER BY j.created_at DESC LIMIT 10"
        ).fetchall()
        conn.close()
        if not rows:
            await q.edit_message_text("No pending BOQ jobs.", reply_markup=admin_menu_kb()); return
        text = "📦 *Pending BOQ Jobs*\n\n"
        for j in rows:
            j = dict(j)
            text += (
                f"• `{j['boq_code']}` | {j.get('buyer_name','—')} "
                f"| {j.get('project_type','—')} | {j.get('status','—')}\n"
                f"  _To deliver result:_ `/deliverboq {j['boq_code']}`\n\n"
            )
        text += "_Use /getfile FILE\\_ID to open any uploaded BOQ file._"
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=admin_menu_kb())

    # ── Verify ────────────────────────────────────────────────────────────────
    elif action == "verify":
        conn = get_connection()
        rows = conn.execute(
            "SELECT * FROM suppliers WHERE verified=0 ORDER BY created_at DESC LIMIT 10"
        ).fetchall()
        conn.close()
        if not rows:
            await q.edit_message_text("No unverified suppliers pending.",
                                       reply_markup=admin_menu_kb()); return
        buttons = []
        for s in rows:
            s = dict(s)
            buttons.append([(
                f"HBH-S-{s['supplier_id']:04d} — {s.get('business_name','?')}",
                f"adm_verify:{s['supplier_id']}:show"
            )])
        buttons.append([("🔙 Back", "adm:home")])
        await q.edit_message_text("🔍 *Pending verifications:*",
                                   parse_mode="Markdown", reply_markup=kb(buttons))

    # ── Prices ────────────────────────────────────────────────────────────────
    elif action == "prices":
        prices = get_all_active_prices()
        if not prices:
            await q.edit_message_text(
                "No prices in database.\n\nUse ➕ Add or 📎 Upload to add prices.",
                reply_markup=price_mgmt_kb([])
            ); return
        await q.edit_message_text(
            f"💰 *Price Management*\n_{len(prices)} active entries_\n\n"
            f"Tap any entry to edit or delete it.",
            parse_mode="Markdown", reply_markup=price_mgmt_kb(prices)
        )

    # ── Categories ────────────────────────────────────────────────────────────
    elif action == "cats":
        cats = get_active_categories()
        await q.edit_message_text(
            "📂 *Category Management*\n\n"
            "Tap any category to toggle active/archived.\n"
            "Categories control what buyers and suppliers can select.",
            parse_mode="Markdown", reply_markup=cat_mgmt_kb(cats)
        )

    # ── Broadcast ─────────────────────────────────────────────────────────────
    elif action == "broadcast":
        await q.edit_message_text(
            "📢 *Broadcast*\n\nWho should receive it?",
            reply_markup=kb([
                [("👥 All Buyers",    "adm_bc:buyers"),
                 ("🏭 All Suppliers", "adm_bc:suppliers")],
                [("🌐 Everyone",      "adm_bc:all")],
                [("🔙 Cancel",        "adm:home")],
            ])
        )


# ── ROUTE PO TO SUPPLIERS (auto-matching) ─────────────────────────────────────

@admin_only
async def route_po_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    po_id   = int(q.data.split(":")[1])
    po      = get_po(po_id)
    buyer   = get_buyer_by_id(po['buyer_id'])

    try:    po_cats = json.loads(po.get('categories', '[]'))
    except: po_cats = []

    suppliers = get_suppliers_matching_categories(po_cats)
    if not suppliers:
        await q.edit_message_text(
            f"No suppliers found for categories in {po['po_code']}.\n"
            f"Add/verify suppliers in those categories first.",
            reply_markup=admin_menu_kb()
        ); return

    notified = 0
    file_delivery_success = 0
    sup_bot = Bot(token=SUPPLIER_BOT_TOKEN)
    
    for idx, s in enumerate(suppliers[:6], 1):          # cap at 6 per PO
        try:
            logger.debug(f"ADMIN ROUTE: Sending PO {po['po_code']} to supplier {s['supplier_id']} (tg:{s['telegram_id']}) - {idx}/6")
            
            # Send text message with PO details
            try:
                await sup_bot.send_message(
                    s['telegram_id'],
                    fmt_po_lead(po, buyer or {}),
                    parse_mode="Markdown",
                    reply_markup=lead_kb(po_id)
                )
                logger.info(f"ADMIN ROUTE: Text sent to supplier {s['supplier_id']}")
            except Exception as e:
                logger.warning(f"ADMIN ROUTE: Failed to send text to supplier {s['supplier_id']}: {e}")
                continue
            
            # Send PO file if it exists
            if po.get('po_file_id'):
                try:
                    logger.debug(f"ADMIN ROUTE: Sending PO file to supplier {s['supplier_id']}")
                    file_caption = (
                        f"📎 *Attachment for {po['po_code']}*\n\n"
                        f"_Buyer: {buyer.get('name','—') if buyer else '—'}_\n"
                        f"_via Habesha Build Hub_ 🏗️"
                    )
                    file_sent = await send_file_preview(
                        bot=sup_bot,
                        chat_id=s['telegram_id'],
                        file_id=po['po_file_id'],
                        filename=po.get('po_file_name', 'po_attachment'),
                        caption=file_caption,
                        parse_mode="Markdown"
                    )
                    if file_sent:
                        logger.info(f"ADMIN ROUTE: File delivered to supplier {s['supplier_id']}")
                        file_delivery_success += 1
                    else:
                        logger.warning(f"ADMIN ROUTE: File delivery failed for supplier {s['supplier_id']}")
                except Exception as e:
                    logger.warning(f"ADMIN ROUTE: Error sending file to supplier {s['supplier_id']}: {e}")
            
            # Increment leads_received
            conn = get_connection()
            conn.execute(
                "UPDATE suppliers SET leads_received=leads_received+1 WHERE supplier_id=?",
                (s['supplier_id'],)
            )
            conn.commit(); conn.close()
            notified += 1
            logger.info(f"ADMIN ROUTE: Successfully routed PO {po['po_code']} to supplier {s['supplier_id']}")
        except Exception as e:
            logger.error(f"ADMIN ROUTE: Unexpected error routing to supplier {s['supplier_id']}: {e}", exc_info=e)

    log_event('po_routed', 'admin', update.effective_user.id, po_id,
              {'suppliers_notified': notified, 'files_delivered': file_delivery_success})
    await q.edit_message_text(
        f"✅ *PO Routed!*\n\n"
        f"`{po['po_code']}` sent to *{notified}* supplier(s)\n"
        f"Files delivered: *{file_delivery_success}*",
        parse_mode="Markdown", reply_markup=admin_menu_kb()
    )


# ── PUSH QUOTES TO BUYER ───────────────────────────────────────────────────────

@admin_only
async def push_quotes_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    po_id  = int(q.data.split(":")[1])
    po     = get_po(po_id)
    quotes = get_po_quotes(po_id)
    buyer  = get_buyer_by_id(po['buyer_id'])
    if not buyer or not quotes:
        await q.edit_message_text("No quotes or buyer not found.",
                                   reply_markup=admin_menu_kb()); return
    try:
        buyer_bot = Bot(token=BUYER_BOT_TOKEN)
        await buyer_bot.send_message(
            buyer['telegram_id'],
            fmt_quotes(quotes, po),
            parse_mode="Markdown",
            reply_markup=quotes_kb(quotes)
        )
        await q.edit_message_text(
            f"✅ Quotes pushed to buyer *{buyer.get('name','?')}*.",
            parse_mode="Markdown", reply_markup=admin_menu_kb()
        )
    except Exception as e:
        await q.edit_message_text(f"Failed: {e}", reply_markup=admin_menu_kb())


# ── SUPPLIER VERIFICATION ──────────────────────────────────────────────────────

@admin_only
async def verify_supplier_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    parts       = q.data.split(":")
    supplier_id = int(parts[1])
    action      = parts[2]
    conn = get_connection()
    s = dict(conn.execute(
        "SELECT * FROM suppliers WHERE supplier_id=?", (supplier_id,)
    ).fetchone())
    conn.close()

    if action == "show":
        await q.edit_message_text(
            f"🔍 *Verification Review*\n\n"
            f"🆔 HBH-S-{s['supplier_id']:04d}\n"
            f"🏢 {s.get('business_name','—')}\n"
            f"📱 {s.get('phone','—')}\n"
            f"📍 {s.get('city','—')}\n"
            f"📅 Joined: {s.get('created_at','')[:10]}\n",
            parse_mode="Markdown", reply_markup=verify_action_kb(supplier_id)
        )

    elif action == "approve":
        conn = get_connection()
        conn.execute("UPDATE suppliers SET verified=1 WHERE supplier_id=?", (supplier_id,))
        conn.commit(); conn.close()
        try:
            sup_bot = Bot(token=SUPPLIER_BOT_TOKEN)
            await sup_bot.send_message(
                s['telegram_id'],
                "✅ *Your account has been verified!*\n\n"
                "Your Verified badge is now active on your profile.\n"
                "Buyers can see you are a trusted supplier on Habesha Build Hub. 🏗️",
                parse_mode="Markdown"
            )
        except Exception: pass
        log_event('supplier_verified', 'admin', update.effective_user.id, supplier_id)
        await q.edit_message_text(
            f"✅ *{s.get('business_name','?')}* verified successfully.",
            reply_markup=admin_menu_kb()
        )

    elif action == "reject":
        try:
            sup_bot = Bot(token=SUPPLIER_BOT_TOKEN)
            await sup_bot.send_message(
                s['telegram_id'],
                "ℹ️ *Verification update*\n\n"
                "We could not verify your account yet.\n"
                "Please send your trade license to @hbh_supportbot to complete verification.",
                parse_mode="Markdown"
            )
        except Exception: pass
        await q.edit_message_text(
            f"Verification rejected — {s.get('business_name','?')} notified.",
            reply_markup=admin_menu_kb()
        )


# ── PRICE MANAGEMENT ───────────────────────────────────────────────────────────

@admin_only
async def price_action_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    parts = q.data.split(":")
    sub   = parts[1]

    if sub == "edit":
        price_id = int(parts[2])
        conn = get_connection()
        p = dict(conn.execute(
            "SELECT * FROM price_reports WHERE price_id=?", (price_id,)
        ).fetchone())
        conn.close()
        context.user_data['editing_price_id'] = price_id
        await q.edit_message_text(
            f"✏️ *Edit Price*\n\n"
            f"Current: *{p['material_type']} {p['brand']}* — "
            f"{p['price_etb']:,.0f} ETB/{p['unit']}\n\n"
            f"Type the new price (numbers only):",
            parse_mode="Markdown"
        )
        return ADM_PRICE_EDIT

    elif sub == "del":
        price_id = int(parts[2])
        delete_price(price_id)
        log_event('price_deleted', 'admin', update.effective_user.id, price_id)
        prices = get_all_active_prices()
        await q.edit_message_text(
            f"🗑️ Price entry deleted.\n\n"
            f"_{len(prices)} active entries remaining_",
            parse_mode="Markdown", reply_markup=price_mgmt_kb(prices)
        )

    elif sub == "add":
        await q.edit_message_text(
            "➕ *Add New Price*\n\n"
            "Type in this format:\n"
            "`MATERIAL | BRAND | UNIT | PRICE`\n\n"
            "Example:\n"
            "`Cement | Derba | bag (50kg) | 435`",
            parse_mode="Markdown"
        )
        return ADM_PRICE_ADD

    elif sub == "upload":
        await q.edit_message_text(
            "📎 *Upload Price File*\n\n"
            "Send a *CSV* or *Excel (.xlsx)* file.\n\n"
            "Required columns (in order):\n"
            "`material_type | brand | unit | price_etb`\n\n"
            "First row must be headers.\n"
            "Example row: `Cement | Derba | bag (50kg) | 430`",
            parse_mode="Markdown"
        )
        context.user_data['awaiting_price_file'] = True
        return ADM_PRICE_EDIT   # reuse state — file handled in text/file handler


@admin_only
async def price_edit_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handles typed new price value OR typed add-price row."""
    text = update.message.text.strip()

    # If we were editing a specific price_id
    if 'editing_price_id' in context.user_data:
        price_id = context.user_data.pop('editing_price_id')
        try:
            new_val = float(text.replace(',', ''))
            edit_price(price_id, new_val)
            log_event('price_edited', 'admin', update.effective_user.id, price_id)
            prices = get_all_active_prices()
            await update.message.reply_text(
                f"✅ Price updated to *{new_val:,.0f} ETB*",
                parse_mode="Markdown", reply_markup=price_mgmt_kb(prices)
            )
        except ValueError:
            await update.message.reply_text("Invalid number. Try again.")
            context.user_data['editing_price_id'] = price_id
            return ADM_PRICE_EDIT
        return ConversationHandler.END

    # If we're adding a new price row via text
    parts = [p.strip() for p in text.split('|')]
    if len(parts) != 4:
        await update.message.reply_text(
            "Format: `MATERIAL | BRAND | UNIT | PRICE`\n\n"
            "Example: `Cement | Derba | bag (50kg) | 430`",
            parse_mode="Markdown"
        )
        return ADM_PRICE_ADD
    try:
        mat, brand, unit, price_str = parts
        price_val = float(price_str.replace(',', ''))
        add_price(mat, brand, unit, price_val,
                  source='Admin', admin_id=update.effective_user.id)
        log_event('price_added', 'admin', update.effective_user.id, 0,
                  {'material': mat, 'brand': brand, 'price': price_val})
        prices = get_all_active_prices()
        await update.message.reply_text(
            f"✅ Added: *{mat} {brand}* — {price_val:,.0f} ETB/{unit}",
            parse_mode="Markdown", reply_markup=price_mgmt_kb(prices)
        )
        return ConversationHandler.END
    except Exception as e:
        await update.message.reply_text(f"Error: {e}\n\nTry again or /cancel")
        return ADM_PRICE_ADD


@admin_only
async def price_file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handles CSV or Excel file upload for bulk price import."""
    doc = update.message.document
    if not doc:
        await update.message.reply_text("Please send a CSV or Excel file."); return

    fname = doc.file_name or ""
    file  = await doc.get_file()
    data  = await file.download_as_bytearray()

    rows   = []
    errors = []

    try:
        if fname.endswith('.csv'):
            text   = data.decode('utf-8', errors='replace')
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                # Normalize keys to lowercase stripped
                rows.append({k.strip().lower(): v for k, v in row.items()})

        elif fname.endswith('.xlsx'):
            import openpyxl
            wb   = openpyxl.load_workbook(io.BytesIO(bytes(data)))
            ws   = wb.active
            hdrs = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows())]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(hdrs, [str(v).strip() if v is not None else '' for v in row])))
        else:
            await update.message.reply_text(
                "Unsupported file type. Please send a .csv or .xlsx file."
            ); return

    except Exception as e:
        await update.message.reply_text(f"Could not parse file: {e}"); return

    # Normalize column names — accept flexible headers
    normalized = []
    for r in rows:
        normalized.append({
            'material_type': r.get('material_type') or r.get('material') or r.get('type',''),
            'brand':         r.get('brand',''),
            'unit':          r.get('unit',''),
            'price_etb':     r.get('price_etb') or r.get('price') or r.get('etb',''),
        })

    result = bulk_import_prices(normalized, admin_id=update.effective_user.id)

    summary = (
        f"📎 *Bulk Import Complete*\n\n"
        f"✅ Imported: {result['imported']}\n"
        f"⏭️ Skipped: {result['skipped']}\n"
        f"🆔 Batch ID: `{result['batch_id']}`\n"
        f"_(Use batch ID to rollback entire upload if needed)_"
    )
    if result['errors']:
        summary += f"\n\n⚠️ *Errors ({len(result['errors'])}):*\n"
        summary += "\n".join(result['errors'][:5])

    prices = get_all_active_prices()
    await update.message.reply_text(
        summary, parse_mode="Markdown", reply_markup=price_mgmt_kb(prices)
    )
    context.user_data.pop('awaiting_price_file', None)
    return ConversationHandler.END


# ── /getfile — admin downloads any file by file ID ────────────────────────────

@admin_only
async def getfile_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /getfile FILE_ID
    Admin previews any file: BOQ uploads, proformas, PO attachments.
    Images shown as photos, documents as downloadable files.
    """
    parts = update.message.text.strip().split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await update.message.reply_text(
            "📎 *How to use /getfile:*\n\n"
            "Type `/getfile` followed by the file ID.\n\n"
            "*Example:*\n"
            "`/getfile BQACAgIAAxkBAAI...`\n\n"
            "*Where to find file IDs:*\n"
            "• BOQ jobs: shown in the 📦 BOQ Queue notification\n"
            "• Proformas: shown in supplier quote notifications\n"
            "• PO attachments: shown in new PO notifications\n\n"
            "_The ID starts with BQAC or AgAC_",
            parse_mode="Markdown",
            reply_markup=admin_menu_kb()
        )
        return

    file_id  = parts[1].strip()
    thinking = await update.message.reply_text("⏳ Fetching file...")
    success  = await send_file_preview(
        bot=context.bot,
        chat_id=update.effective_chat.id,
        file_id=file_id,
        filename=context.user_data.get(f'filename_{file_id}', 'hbh_file'),
        caption=f"📎 File retrieved via HBH Admin.\n`{file_id}`",
        parse_mode="Markdown"
    )
    try:
        await thinking.delete()
    except Exception:
        pass
    if not success:
        await update.message.reply_text(
            "❌ Could not fetch that file.\n\n"
            "Check that you copied the full file ID correctly.",
            reply_markup=admin_menu_kb()
        )


# ── BOQ RESULT DELIVERY — admin sends completed BOQ back to buyer ──────────────

ADM_BOQ_DELIVER_FILE = 80
ADM_BOQ_DELIVER_NOTE = 81

@admin_only
async def boq_deliver_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    /deliverboq BOQ_CODE
    Admin starts the flow to send a completed BOQ result back to the buyer.
    """
    parts = update.message.text.strip().split(maxsplit=1)
    if len(parts) < 2:
        await update.message.reply_text(
            "📊 *Deliver BOQ Result to Buyer*\n\n"
            "Usage: `/deliverboq BOQ_CODE`\n\n"
            "Example: `/deliverboq BOQ-12345`\n\n"
            "_Find the BOQ code in 📦 BOQ Queue in your admin menu._",
            parse_mode="Markdown",
            reply_markup=admin_menu_kb()
        )
        return

    boq_code = parts[1].strip().upper()
    conn = get_connection()
    job  = conn.execute(
        "SELECT j.*, b.name as buyer_name, b.telegram_id as buyer_tg "
        "FROM boq_jobs j JOIN buyers b ON j.buyer_id=b.buyer_id "
        "WHERE j.boq_code=?", (boq_code,)
    ).fetchone()
    conn.close()

    if not job:
        await update.message.reply_text(
            f"❌ BOQ code `{boq_code}` not found.\n\n"
            "Check the code and try again.",
            parse_mode="Markdown",
            reply_markup=admin_menu_kb()
        )
        return

    job = dict(job)
    context.user_data['boq_deliver'] = {
        'boq_code':  job['boq_code'],
        'buyer_tg':  job['buyer_tg'],
        'buyer_name': job['buyer_name'],
        'scope':     job.get('scope','—'),
    }

    await update.message.reply_text(
        f"📊 *Delivering BOQ result*\n\n"
        f"🆔 Ticket: `{job['boq_code']}`\n"
        f"👤 Buyer: {job['buyer_name']}\n"
        f"📋 Scope: {job.get('scope','—')}\n\n"
        f"*Send the completed BOQ file now:*\n"
        f"_Excel (.xlsx), PDF, or any file format_",
        parse_mode="Markdown"
    )
    return ADM_BOQ_DELIVER_FILE


@admin_only
async def boq_deliver_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin uploads the completed BOQ file."""
    doc = update.message.document or (
        update.message.photo[-1] if update.message.photo else None
    )
    if not doc:
        await update.message.reply_text(
            "Please send the BOQ result file. Or type /cancel to stop."
        )
        return ADM_BOQ_DELIVER_FILE

    context.user_data['boq_deliver']['file_id']   = doc.file_id
    context.user_data['boq_deliver']['file_name'] = getattr(doc, 'file_name', 'boq_result.xlsx')

    await update.message.reply_text(
        f"✅ File received: *{context.user_data['boq_deliver']['file_name']}*\n\n"
        f"Add a note for the buyer?\n"
        f"_e.g. 'Total estimated cost: 2.4M ETB — see breakdown inside'\n"
        f"or type 'none' to skip_",
        parse_mode="Markdown"
    )
    return ADM_BOQ_DELIVER_NOTE


@admin_only
async def boq_deliver_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin types a note, then the system delivers to buyer."""
    note = update.message.text.strip()
    if note.lower() == 'none':
        note = ''

    bd = context.user_data.pop('boq_deliver', {})

    # Mark BOQ job as delivered in database
    conn = get_connection()
    conn.execute(
        "UPDATE boq_jobs SET status='delivered', notes=? WHERE boq_code=?",
        (note, bd.get('boq_code',''))
    )
    conn.commit()
    conn.close()

    # Send the file directly to the buyer via the buyer bot
    try:
        buyer_bot = Bot(token=BUYER_BOT_TOKEN)
        file_id   = bd.get('file_id','')
        caption   = (
            f"📊 *Your BOQ Estimate is Ready!*\n\n"
            f"🆔 Ticket: `{bd.get('boq_code','')}`\n"
            f"📋 Scope: {bd.get('scope','—')}\n"
        )
        if note:
            caption += f"\n💬 *Note from specialist:*\n{note}\n"
        caption += (
            f"\n📎 *File ID for future access:*\n"
            f"`{file_id}`\n"
            f"_Save this ID. Use `/getfile {file_id}` anytime to re-download._\n\n"
            f"_Delivered by Habesha Build Hub specialists_ 🏗️"
        )
        await send_file_preview(
            bot=buyer_bot,
            chat_id=bd['buyer_tg'],
            file_id=file_id,
            filename=bd.get('file_name', 'boq_result.xlsx'),
            caption=caption,
            parse_mode="Markdown"
        )
        log_event('boq_delivered', 'admin', update.effective_user.id, 0,
                  {'boq_code': bd.get('boq_code')})
        await update.message.reply_text(
            f"✅ *BOQ result delivered!*\n\n"
            f"Ticket `{bd.get('boq_code','')}` sent to *{bd.get('buyer_name','?')}*.\n"
            f"They can re-download it anytime with `/getfile {file_id}`",
            parse_mode="Markdown",
            reply_markup=admin_menu_kb()
        )
    except Exception as e:
        logger.error(f"BOQ delivery failed: {e}")
        await update.message.reply_text(
            f"❌ Delivery failed: {e}\n\n"
            f"The buyer's Telegram ID may be invalid or they have blocked the bot.",
            reply_markup=admin_menu_kb()
        )
    return ConversationHandler.END

@admin_only
async def handle_rollback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Usage: /rollback BATCH_ID"""
    parts = update.message.text.strip().split()
    if len(parts) != 2:
        await update.message.reply_text("Usage: /rollback BATCH_ID"); return
    bid = parts[1]
    rollback_batch(bid)
    log_event('price_batch_rolled_back', 'admin', update.effective_user.id, 0, {'batch_id': bid})
    prices = get_all_active_prices()
    await update.message.reply_text(
        f"✅ Batch `{bid}` rolled back. {len(prices)} active prices remain.",
        parse_mode="Markdown", reply_markup=price_mgmt_kb(prices)
    )


# ── CATEGORY MANAGEMENT ────────────────────────────────────────────────────────

@admin_only
async def cat_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    parts = q.data.split(":")
    sub   = parts[1]

    if sub == "toggle":
        cat_id = int(parts[2])
        conn   = get_connection()
        cur    = conn.execute(
            "SELECT is_active FROM categories WHERE cat_id=?", (cat_id,)
        ).fetchone()
        if cur:
            new_val = 0 if cur[0] else 1
            conn.execute("UPDATE categories SET is_active=? WHERE cat_id=?", (new_val, cat_id))
            conn.commit()
        conn.close()
        cats = get_active_categories()
        await q.edit_message_reply_markup(reply_markup=cat_mgmt_kb(cats))

    elif sub == "add":
        await q.edit_message_text(
            "➕ *Add Category*\n\n"
            "Type in format:\n`KEY | Label | Emoji`\n\n"
            "Example: `glass | 🪟 Glass & Glazing | 🪟`",
            parse_mode="Markdown"
        )
        return ADM_CAT_ADD


@admin_only
async def cat_add_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text  = update.message.text.strip()
    parts = [p.strip() for p in text.split('|')]
    if len(parts) < 2:
        await update.message.reply_text(
            "Format: `KEY | Label | Emoji`\nExample: `glass | Glass & Glazing | 🪟`",
            parse_mode="Markdown"
        ); return ADM_CAT_ADD
    key    = parts[0].lower().replace(' ', '_')
    label  = parts[1]
    emoji  = parts[2] if len(parts) > 2 else '📦'
    ok = add_category(key, label, emoji)
    if ok:
        await update.message.reply_text(
            f"✅ Category *{label}* added.",
            parse_mode="Markdown", reply_markup=admin_menu_kb()
        )
    else:
        await update.message.reply_text(
            f"Category key `{key}` already exists. Use a different key.",
            parse_mode="Markdown"
        )
    return ConversationHandler.END


# ── BROADCAST ──────────────────────────────────────────────────────────────────

@admin_only
async def bc_target_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    context.user_data['bc_target'] = q.data.split(":")[1]
    await q.edit_message_text(
        f"📢 Broadcast to: *{context.user_data['bc_target']}*\n\n"
        f"Type the message to send:",
        parse_mode="Markdown"
    )
    return ADM_BROADCAST_TEXT


@admin_only
async def bc_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text   = update.message.text.strip()
    target = context.user_data.pop('bc_target', 'all')
    conn   = get_connection()

    total  = 0
    failed = 0

    if target in ('buyers', 'all'):
        buyer_bot = Bot(token=BUYER_BOT_TOKEN)
        for (tid,) in conn.execute("SELECT telegram_id FROM buyers").fetchall():
            try:
                await buyer_bot.send_message(tid, text, parse_mode="Markdown")
                total += 1
            except Exception:
                failed += 1

    if target in ('suppliers', 'all'):
        sup_bot = Bot(token=SUPPLIER_BOT_TOKEN)
        for (tid,) in conn.execute("SELECT telegram_id FROM suppliers").fetchall():
            try:
                await sup_bot.send_message(tid, text, parse_mode="Markdown")
                total += 1
            except Exception:
                failed += 1

    conn.close()
    await update.message.reply_text(
        f"✅ Broadcast sent to *{total}* users. Failed: {failed}",
        parse_mode="Markdown", reply_markup=admin_menu_kb()
    )
    return ConversationHandler.END


# ── GENERAL TEXT HANDLER (handles awaiting states without conversation) ─────────

@admin_only
async def general_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Use /start to open the admin menu.",
        reply_markup=admin_menu_kb()
    )


async def error_handler(update, context: ContextTypes.DEFAULT_TYPE):
    logger.error(f"Error: {context.error}", exc_info=context.error)


# ── BUILD ──────────────────────────────────────────────────────────────────────

def build_admin_app():
    app = Application.builder().token(ADMIN_BOT_TOKEN).build()

    price_conv = ConversationHandler(
        per_chat=True,
        per_user=True,
        per_message=True,
        entry_points=[
            CallbackQueryHandler(price_action_callback, pattern=r"^adm_price:(edit|add|upload):?"),
        ],
        states={
            ADM_PRICE_EDIT: [
                MessageHandler(filters.Document.ALL, price_file_handler),
                MessageHandler(filters.TEXT & ~filters.COMMAND, price_edit_text),
            ],
            ADM_PRICE_ADD: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, price_edit_text),
            ],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        allow_reentry=True,
    )

    cat_conv = ConversationHandler(
        per_chat=True,
        per_user=True,
        per_message=True,
        entry_points=[CallbackQueryHandler(cat_callback, pattern=r"^adm_cat:add")],
        states={ADM_CAT_ADD: [MessageHandler(filters.TEXT & ~filters.COMMAND, cat_add_text)]},
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        allow_reentry=True,
    )

    bc_conv = ConversationHandler(
        per_chat=True,
        per_user=True,
        per_message=True,
        entry_points=[CallbackQueryHandler(bc_target_callback, pattern=r"^adm_bc:")],
        states={ADM_BROADCAST_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, bc_send)]},
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
    )

    boq_deliver_conv = ConversationHandler(
        per_chat=True,
        per_user=True,
        per_message=True,
        entry_points=[CommandHandler("deliverboq", boq_deliver_start)],
        states={
            ADM_BOQ_DELIVER_FILE: [
                MessageHandler(
                    (filters.Document.ALL | filters.PHOTO) & ~filters.COMMAND,
                    boq_deliver_file
                ),
            ],
            ADM_BOQ_DELIVER_NOTE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, boq_deliver_note),
            ],
        },
        fallbacks=[CommandHandler("cancel", lambda u, c: ConversationHandler.END)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start",      start))
    app.add_handler(CommandHandler("rollback",   handle_rollback))
    app.add_handler(CommandHandler("getfile",    getfile_cmd))
    app.add_handler(boq_deliver_conv)
    app.add_handler(price_conv)
    app.add_handler(cat_conv)
    app.add_handler(bc_conv)
    app.add_handler(CallbackQueryHandler(adm_callback,            pattern=r"^adm:"))
    app.add_handler(CallbackQueryHandler(route_po_callback,       pattern=r"^adm_route:"))
    app.add_handler(CallbackQueryHandler(push_quotes_callback,    pattern=r"^adm_pushquotes:"))
    app.add_handler(CallbackQueryHandler(verify_supplier_callback,pattern=r"^adm_verify:"))
    app.add_handler(CallbackQueryHandler(price_action_callback,   pattern=r"^adm_price:del:"))
    app.add_handler(CallbackQueryHandler(cat_callback,            pattern=r"^adm_cat:toggle:"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, general_text))
    app.add_error_handler(error_handler)
    return app
